"""Starlette ASGI server for the App2 (HTMX) dashboard renderer.

X.2.b shape — all-GET REST surface (no POSTs except dev-log):

- ``GET  /`` — 302 redirect to ``/dashboards``. The dashboards
  list IS the canonical entry; ``/`` is convenience.
- ``GET  /dashboards`` — landing page listing every dashboard the
  server is wired to serve. One link per dashboard, bookmarkable
  per entry.
- ``GET  /dashboards/{dashboard_id}`` — dashboard chrome + the
  served Sheet inline. 404 if the dashboard_id isn't in the
  wired ``dashboards`` mapping.
- ``GET  /dashboards/{dashboard_id}/sheets/{sheet_id}/visuals/{visual_id}/data``
  — chart data fragment for HTMX swap. Filter values arrive as
  query string. GET-not-POST means every (visual, filter-set)
  tuple is a bookmarkable URL.
- ``POST /log`` (dev-only, gated by ``dev_log=True``) — the only
  POST route. Receives forwarded HTMX + d3 click events from the
  browser for live debugging.

X.2.b.3: ``make_app`` takes a ``dashboards`` mapping so one server
can host multiple apps. Each value is a ``ServedDashboard`` carrying
its own tree, sheet, title, and data fetcher (different apps query
different matviews via different fetchers). X.2.g wires the four
QS apps (Executives / Investigation / L2 Flow Tracing / L1
Dashboard) into this mapping from one L2 instance.

Error handling (X.2.m)
----------------------

Two exception handlers wrap the app so production deploys never
return a Starlette default error page:

- ``HTTPException(404)`` (raised by route handlers when a
  dashboard_id / sheet_id slug doesn't resolve) renders a themed
  "Not found" page via ``emit_error_page``.
- Generic ``Exception`` (anything uncaught from a route handler —
  fetcher SQL crash, render-time bug, DB unreachable) returns 500
  with a themed "Something went wrong" page. ``dev_log=True``
  carries the traceback inside a collapsible ``<details>``;
  production hides it.

The HTMX ``htmx:responseError`` event in ``bootstrap.js`` surfaces
a transient toast for 4xx / 5xx responses to swap targets so a
failed visual data fetch shows context instead of a blank panel.

Pluggable data fetcher
----------------------

Each ``ServedDashboard`` owns a ``DataFetcher`` callable so the
spike + tests can run without a database:

    def stub(visual_id: VisualId, params: Mapping[str, list[str]]) -> Any:
        return {"nodes": [...], "links": [...]}

    app = make_app(dashboards={
        "smoke": ServedDashboard(
            tree_app=app, sheet=money_trail,
            title="Smoke", data_fetcher=stub,
        ),
    })

Production deploys wire the same callable to a DB-backed factory
(see ``_db_fetcher.make_db_fetcher``).

Stateless on purpose
--------------------

No sessions, no auth, no in-process caching. Each GET executes the
fetcher fresh. Cache-Control headers (X.2.b.4) push caching to
edge / browser layers — the URL IS the cache key, by design.
"""

from __future__ import annotations

import html as _html_module
import inspect
import json
import logging
import re
import traceback

html_escape = _html_module.escape
from collections.abc import Awaitable, Callable, Mapping, Sequence
from dataclasses import dataclass, replace
from typing import Any, Union

from pathlib import Path

from starlette.applications import Starlette
from starlette.concurrency import run_in_threadpool
from starlette.datastructures import QueryParams
from starlette.exceptions import HTTPException
from starlette.requests import Request
from starlette.responses import (
    HTMLResponse, JSONResponse, RedirectResponse, Response,
)
from starlette.routing import Mount, Route
from starlette.staticfiles import StaticFiles

# Y.2.gate.c.11.app2-server-logs — dedicated logger for browser-
# forwarded dev events (POST /log payloads). Routed through Python's
# logging module so the test harness's per-run log capture
# (`tests/e2e/_harness_html2.py::_attach_app2_log_handler`) lands
# the events alongside uvicorn's access log under
# ``$RECON_GEN_RUN_DIR/app2/server.log``.
_DEVLOG = logging.getLogger("recon_gen.app2.devlog")


from recon_gen.common.db import PoolReleasedDuringRefresh
from recon_gen.common.html._side_panel import metadata_panel_route_factory
from recon_gen.common.html._tree_fetcher import OptionsSearchFetcher
from recon_gen.common.html._tree_filter_specs import (
    make_filter_specs_for_sheet,
)
from recon_gen.common.html.render import (
    FilterSpec,
    ParameterDateSpec,
    ParameterDropdownSpec,
    ParameterMultiSelectSpec,
    ParameterNumberSpec,
    emit_dashboards_list,
    emit_error_page,
    emit_html,
    emit_visual_data_fragment,
)
from recon_gen.common.ids import VisualId
from recon_gen.common.l2.theme import ThemePreset
from recon_gen.common.tree.structure import App, Sheet


# (visual_id, filter_params) → chart data shaped for the visual's
# d3 hydrator. The renderer just JSON-serializes whatever the
# fetcher returns; the per-visual shape contract lives in the
# bootstrap.js renderXxx functions.
#
# Two shapes accepted (X.2.n.5):
#   - async: production fetcher built by make_tree_db_fetcher.
#     ``visual_data`` route awaits it directly.
#   - sync: stub fetchers in tests + the legacy _db_fetcher path.
#     ``visual_data`` wraps them in run_in_threadpool so they
#     don't block the event loop.
# ``inspect.iscoroutinefunction`` picks the dispatch at request time.
# X.2.o.3: ``VisualId`` not ``str`` so the fetcher contract ties
# back to the tree's typed visual identifier. Test stubs typed as
# ``Callable[[str, ...], ...]`` remain assignable here via Callable
# parameter contravariance (str is wider than VisualId on input).
DataFetcher = Union[
    Callable[[VisualId, Mapping[str, list[str]]], Awaitable[Any]],
    Callable[[VisualId, Mapping[str, list[str]]], Any],
]


@dataclass(frozen=True)
class ServedDashboard:
    """One dashboard's wiring for the App2 server.

    Each App2 server holds a mapping ``{dashboard_id: ServedDashboard}``
    so one process can serve multiple apps from one L2 instance
    (X.2.g wires Executives + Investigation + L2FT + L1 from one
    L2). Per-dashboard fetcher means apps that query different
    matviews don't have to share a routing layer.

    Attributes:
        tree_app: tree ``App`` node owning the analysis the sheet
            lives in. Internal IDs are resolved on first emit
            (idempotent).
        sheet: tree ``Sheet`` rendered at ``/dashboards/{id}``. Must
            belong to ``tree_app.analysis.sheets``.
        title: human-readable name for the ``/dashboards`` listing.
        data_fetcher: per-dashboard fetcher invoked on every GET to
            the visual data path. Returns d3-shaped chart data.
        theme: per-dashboard ``ThemePreset`` injected as CSS
            variables in the page shell. ``None`` falls back to
            ``DEFAULT_PRESET`` (silent-fallback per N.4.k, mirrors
            QS dialect's CLASSIC fallback). Multi-dashboard servers
            usually share a single theme since the listing page
            renders one palette across all entries.
    """
    tree_app: App
    sheet: Sheet
    title: str
    data_fetcher: DataFetcher
    theme: ThemePreset | None = None
    # Explicit filter-form specs. ``()`` (the common case for tree-built
    # apps) means the routes auto-derive per-sheet from the tree's
    # parameter-control nodes via ``make_filter_specs_for_sheet`` —
    # currently one ``ParameterMultiSelectSpec`` per MULTI_SELECT
    # ``ParameterDropdown`` (Y.2.app2.cde.l2ft-wiring.b). Supply a
    # non-empty tuple to override (the smoke app does this for its
    # hand-crafted demo filters).
    filter_specs: tuple[FilterSpec, ...] = ()
    # CQ.2 — server-side typeahead fetcher. Drives:
    # - The JSON typeahead endpoint (``dropdown-search/...``) —
    #   Tom Select's ``load`` callback fires ``?q=<typed>`` per
    #   keystroke (debounced 300ms via loadThrottle).
    # - The HTML cascade endpoint (``dropdown-options/...``) —
    #   sibling-control change triggers ``hx-get`` against the
    #   narrowed seed page (query=''); HTMX swaps the inner
    #   <option> set; Tom Select re-wires with empty
    #   loadedSearches → next focus re-runs preload.
    # ``None`` (stub-fetcher tests, which don't carry LinkedValues
    # dropdowns) → those dropdowns render empty + Tom Select
    # preload no-ops (degraded, not a crash).
    options_search_fetcher: OptionsSearchFetcher | None = None


# CR.1 (2026-06-08) — Excel sheet name limit: ≤ 31 chars + no `: \ / ? * [ ]`.
# Pre-CR.1 the export silently truncated `visual_id[:31]`; auto-derived
# UUIDs (36 chars) lost their last 5 hex — vanishing into ~16M collision
# space when several visuals share a workbook. Human-readable visual_ids
# ≤31 (including the 31-char exec-* ones sitting at the wall) pass
# through unchanged; UUIDs collapse to a stable `<first8>-<last4>`
# shortform; >31-char human strings raise so the rename surfaces at
# request time. Unit-tier lint at tests/unit/test_xlsx_sheet_title.py
# walks every app's emitted tree and asserts the contract pre-deploy.
_XLSX_FORBIDDEN_RE = re.compile(r"[:\\/?*\[\]]")
_VISUAL_ID_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
)
XLSX_SHEET_NAME_MAX = 31


def sheet_title_from_visual_id(visual_id: VisualId) -> str:
    """Derive an Excel-safe worksheet title from a ``visual_id``.

    Excel caps sheet names at ``XLSX_SHEET_NAME_MAX`` (31) characters
    and rejects ``: \\ / ? * [ ]``. ``openpyxl`` will silently
    accept >31-char names + write them to the .xlsx; Excel enforces
    at file-open, so two visuals diverging after char 31 silently
    collide. CR.1 closes the silent-truncation hole at the export
    boundary:

    - **UUIDv5 visual_ids** (auto-derived in
      ``common/tree/structure.py::auto_id``, always 36 chars) →
      collapse to a stable 13-char ``<first8>-<last4>`` shortform.
      Deterministic, 48-bit collision space (281T pairs).
    - **Human-readable ≤ 31 chars** → pass through, forbidden chars
      replaced with ``_``.
    - **Human-readable > 31 chars** → raise ``ValueError`` so the
      caller surfaces the rename requirement. The unit-tier lint
      catches this pre-deploy; this raise is the defense-in-depth
      that prevents a silently-truncated worksheet from ever
      reaching disk.
    """
    if _VISUAL_ID_UUID_RE.match(visual_id):
        return f"{visual_id[:8]}-{visual_id[-4:]}"
    sanitized = _XLSX_FORBIDDEN_RE.sub("_", visual_id)
    if len(sanitized) > XLSX_SHEET_NAME_MAX:
        raise ValueError(
            f"visual_id {visual_id!r} ({len(visual_id)} chars) "
            f"exceeds Excel's {XLSX_SHEET_NAME_MAX}-char sheet "
            f"name limit. Rename the visual to ≤ "
            f"{XLSX_SHEET_NAME_MAX} chars."
        )
    return sanitized


def _emit_xlsx_workbook(data: Any, visual_id: VisualId) -> bytes:  # typing-smell: ignore[explicit-any]: visual data is heterogeneous shape_table payload from the route layer
    """CH.5 (2026-06-08) — render a `shape_table`-shaped data dict
    as an XLSX file.

    Reads `data["columns"]` + `data["rows"]` per the contract in
    `_data_shape.py::shape_table` — list of `{"name", "label"?,
    "format"?}` dicts + list-of-list rows. Header row uses `label
    || name`; currency-format columns get Excel currency format
    (right-aligned, $#,##0.00 negatives in red) and feed the rows'
    raw Decimal/float values directly so Excel sums work.

    Raises HTTPException(400) when `data` isn't table-shaped (e.g.
    KPI / Bar / Sankey visuals) — XLSX export only makes sense for
    tables.
    """
    if not isinstance(data, Mapping):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Visual {visual_id!r}: ?format=xlsx requires a "
                f"table-shaped data payload, got "
                f"{type(data).__name__}"
            ),
        )
    columns = data.get("columns")  # pyright: ignore[reportUnknownMemberType, reportUnknownVariableType]: Mapping[str, Any] returns Any
    rows = data.get("rows")  # pyright: ignore[reportUnknownMemberType, reportUnknownVariableType]: Mapping[str, Any] returns Any
    if not isinstance(columns, list) or not isinstance(rows, list):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Visual {visual_id!r}: ?format=xlsx requires "
                f"`columns` + `rows` keys (got "
                f"{sorted(data.keys())!r})"  # pyright: ignore[reportUnknownArgumentType]: data is Mapping[str, Any]; .keys() typed
            ),
        )
    from io import BytesIO  # noqa: PLC0415 — lazy
    from openpyxl import Workbook  # type: ignore[import-untyped]  # noqa: PLC0415 — lazy: only on xlsx click
    from openpyxl.styles import Alignment, Font  # type: ignore[import-untyped]  # noqa: PLC0415 — lazy
    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover — defensive: openpyxl always returns one
        raise HTTPException(
            status_code=500,
            detail="openpyxl Workbook() returned no active sheet",
        )
    # CR.1 — Excel sheet name ≤ 31 chars + forbidden-char set. The
    # pre-CR.1 ``visual_id[:31]`` silently truncated 36-char UUIDv5
    # auto-derived ids + any human-readable id authored >31 chars,
    # leaving the operator to discover collisions at file-open time.
    # ``sheet_title_from_visual_id`` (a) collapses UUIDs to a stable
    # 13-char shortform and (b) raises ``ValueError`` on long human
    # strings so the rename surfaces at request time instead of
    # silently colliding. See the helper's docstring for the full
    # contract.
    try:
        ws.title = sheet_title_from_visual_id(visual_id)
    except ValueError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    # Header row. CY.4.1 — track which positional column indices are
    # hidden from the rendered table; XLSX export drops the same set
    # so the spreadsheet matches what the operator sees on-screen.
    # The row tuple is still positional, so we filter cells by index
    # in lockstep with the column filter.
    headers: list[str] = []
    formats: list[str] = []
    visible_idx: list[int] = []
    for col_idx, col in enumerate(columns):  # pyright: ignore[reportUnknownVariableType, reportUnknownArgumentType]: list element type unknown
        if not isinstance(col, Mapping):
            continue
        if col.get("hidden"):  # pyright: ignore[reportUnknownMemberType, reportUnknownArgumentType]: col is Mapping[str, Any] from shape_table
            continue
        name = str(col.get("name") or "")  # pyright: ignore[reportUnknownMemberType, reportUnknownArgumentType]: col is Mapping[str, Any] from shape_table
        label = str(col.get("label") or name)  # pyright: ignore[reportUnknownMemberType, reportUnknownArgumentType]: col is Mapping[str, Any] from shape_table
        fmt = str(col.get("format") or "")  # pyright: ignore[reportUnknownMemberType, reportUnknownArgumentType]: col is Mapping[str, Any] from shape_table
        headers.append(label)
        formats.append(fmt)
        visible_idx.append(col_idx)
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    # Data rows.
    right_align = Alignment(horizontal="right")
    for row in rows:  # pyright: ignore[reportUnknownVariableType]: list element type unknown
        if not isinstance(row, list):
            continue
        out_row: list[object] = []
        for src_idx in visible_idx:
            if src_idx >= len(row):  # pyright: ignore[reportUnknownArgumentType]: row is list[Unknown] (heterogeneous DB row values)
                continue
            value = row[src_idx]  # pyright: ignore[reportUnknownVariableType]: heterogeneous DB row values
            # openpyxl handles str / int / float / Decimal / bool /
            # None natively. Anything else gets str()-coerced so we
            # don't crash on dates/UUIDs/etc.
            if value is None or isinstance(value, (str, int, float, bool)):
                out_row.append(value)
            else:
                # Decimal in particular passes through cleanly —
                # openpyxl writes it as a numeric cell.
                from decimal import Decimal  # noqa: PLC0415 — lazy
                if isinstance(value, Decimal):
                    out_row.append(value)
                else:
                    out_row.append(str(value))  # pyright: ignore[reportUnknownArgumentType]: heterogeneous DB row values are Any-typed; str() coerces any non-numeric value safely
        ws.append(out_row)
    # Apply currency format + right-align to currency columns.
    for col_idx, fmt in enumerate(formats, start=1):
        if fmt == "currency":
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = (
                    '"$"#,##0.00_);[Red]("$"#,##0.00)'
                )
                cell.alignment = right_align
        elif fmt == "number":
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = "#,##0"
                cell.alignment = right_align
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _query_params_as_multidict(
    query_params: QueryParams,
) -> dict[str, list[str]]:
    """BN.1 — Starlette ``QueryParams`` → the
    ``Mapping[str, list[str]]`` shape the sql-executor + options
    fetcher expect (a repeated ``?param_x=a&param_x=b`` keeps both
    values). Same conversion ``_apply_url_param_overrides`` does
    locally; lifted to a shared helper now that two callers need it."""
    by_name: dict[str, list[str]] = {}
    for key, value in query_params.multi_items():
        by_name.setdefault(key, []).append(value)
    return by_name


def _resolve_linked_options(
    specs: tuple[FilterSpec, ...],
) -> tuple[FilterSpec, ...]:
    """CQ.2.c — pre-CQ.2 this ran one ``SELECT DISTINCT`` per
    LinkedValues spec at sheet render to bake the option universe
    into the ``<select>`` (capped at the silent ``_OPTIONS_CAP = 2000``
    that operator condemned 2026-06-08). Post-CQ.2 the render emits
    LinkedValues pickers with empty ``<option>`` lists + the sticky
    selected value; bootstrap.js wires Tom Select's ``load`` callback
    to the new JSON typeahead endpoint, which serves the seed page
    on first focus and per-keystroke searches after that.

    Kept as an identity passthrough so call sites stay legible during
    the CQ.2 transition; it deletes once tests / cascade path migrate
    in CQ.2.e.
    """
    return tuple(specs)


def _apply_url_param_overrides(
    specs: tuple[FilterSpec, ...],
    query_params: QueryParams,
) -> tuple[FilterSpec, ...]:
    """Pre-fill each parameter-bound filter spec from the page URL's
    ``?param_<name>=<v>`` keys (u.4.e.4 — cross-sheet drills + bookmarkable
    filter state).

    Every visual section loads via ``hx-get`` + ``hx-include="#filter-form"``,
    so the form's *current DOM state* at fire time is what narrows the first
    fetch. When a sheet is opened with ``?param_<name>=<v>`` — a row drill
    that walked an anchor onto a new sheet, or a bookmarked URL — pre-marking
    the matching ``<option>`` (dropdowns) / ``<input>`` value (numeric
    sliders) makes that initial fetch already carry the value, so the
    destination renders narrowed instead of needing a manual re-pick.

    - ``ParameterDropdownSpec`` → ``selected`` = the (last) URL value for
      its name.
    - ``ParameterMultiSelectSpec`` → ``selected`` = every URL value for its
      name (repeated key → tuple).
    - ``ParameterNumberSpec`` → ``default`` = the (last) URL value parsed as
      float; un-parseable input is ignored (the spec keeps its declared
      default).

    Specs whose name has no matching ``param_<name>`` key — and the
    non-parameter spec kinds (``CategoryFilterSpec`` / ``NumericRangeSpec``,
    which key off ``filter_<col>`` / ``min_<col>`` / ``max_<col>``, not
    ``param_<name>``) — pass through unchanged.
    """
    by_name: dict[str, list[str]] = {}
    for key, value in query_params.multi_items():
        if key.startswith("param_"):
            by_name.setdefault(key[len("param_"):], []).append(value)
    if not by_name:
        return specs
    out: list[FilterSpec] = []
    for spec in specs:
        if isinstance(spec, ParameterDropdownSpec):
            vals = by_name.get(spec.name)
            out.append(replace(spec, selected=vals[-1]) if vals else spec)
        elif isinstance(spec, ParameterMultiSelectSpec):
            vals = by_name.get(spec.name)
            out.append(replace(spec, selected=tuple(vals)) if vals else spec)
        elif isinstance(spec, ParameterDateSpec):
            vals = by_name.get(spec.name)
            out.append(replace(spec, selected=vals[-1]) if vals else spec)
        elif isinstance(spec, ParameterNumberSpec):
            vals = by_name.get(spec.name)
            if vals:
                try:
                    out.append(replace(spec, default=float(vals[-1])))
                except ValueError:
                    out.append(spec)
            else:
                out.append(spec)
        else:
            out.append(spec)
    return tuple(out)


def make_app(
    *,
    dashboards: Mapping[str, ServedDashboard],
    dev_log: bool = False,
    visual_data_cache_max_age_s: int = 60,
    docs_dir: Path | None = None,
    studio_routes: Sequence[Route | Mount] | None = None,
    banner_text: str | None = None,
) -> Starlette:
    """Build a Starlette ASGI app serving multiple dashboards.

    Args:
        dashboards: ``{dashboard_id: ServedDashboard}`` mapping.
            One entry per dashboard. The server validates inbound
            path slugs against this mapping; unknown ids 404.
        docs_dir: when set, the *built* mkdocs handbook site at this
            path is mounted at ``/docs`` (``StaticFiles(html=True)`` —
            ``/docs/handbook/l1/`` → ``…/index.html``) and the
            dashboards-list page links it (X.2.i — "one process, one
            place": docs + dashboards together). ``None`` (default)
            leaves ``/docs`` unmounted. The standalone ``docs apply`` /
            ``docs serve`` / ``docs export`` CLI is unchanged either
            way — embedding here is purely additive.
        dev_log: when True, the page emits a ``<meta
            name="dev-log">`` tag that activates the client-side
            event forwarder + a ``POST /log`` route is registered
            that prints each forwarded event to stderr. Off by
            default — keeps production deploys silent and zero-
            overhead. The developer tool / smoke server enables it.
            Also sets ``Cache-Control: no-store`` on visual data
            responses so dev iteration sees fresh data on every
            reload (no surprise stale fragments).
        visual_data_cache_max_age_s: ``Cache-Control: public,
            max-age=N`` on visual-data responses. URL == cache
            key (X.2.b's GET-shape contract), so any (visual,
            filter-set) tuple stays cacheable for ``N`` seconds
            at the edge / browser. Conservative default of 60s
            since matviews refresh on ETL cycles (minutes-to-
            hours); production can dial up. Ignored when
            ``dev_log=True`` (cache is bypassed for dev runs).
        studio_routes: X.4.a.4 — when set, splice these routes in
            BEFORE the Dashboards routes, AND skip the default
            ``GET / → /dashboards`` redirect (since the Studio mount
            owns ``GET /`` as its landing page). ``None`` is the
            Dashboards-only mount: keep the redirect, no Studio
            routes. ``cli.dashboards`` always passes ``None``;
            ``cli.studio`` passes the list ``make_studio_routes``
            built. Severability contract: ``make_app`` does not
            import ``_studio_routes`` — Studio routes are constructed
            externally and injected.

    Returns:
        A ``starlette.Starlette`` ASGI application.
    """
    if not dashboards:
        raise ValueError(
            "make_app requires at least one dashboard in the "
            "`dashboards` mapping."
        )

    # Cache header is the same string on every visual-data
    # response — pre-compute it once instead of formatting per
    # request.
    if dev_log:
        # Dev runs bypass the cache so the developer sees fresh
        # data when reloading a swap. Cache-Control: no-store
        # tells every layer (browser, edge, intermediate proxy)
        # not to keep the response.
        visual_data_cache_header = "no-store"
    else:
        visual_data_cache_header = (
            f"public, max-age={visual_data_cache_max_age_s}"
        )

    # X.2.e — every analysis-attached sheet is reachable as a tab.
    # Snapshot the {dashboard_id: {sheet_id: Sheet}} mapping so the
    # /sheets/:s route can resolve a sheet without walking the tree
    # on every request, and so the 404 path is fast (dict lookup).
    all_sheets: dict[str, dict[str, "Sheet"]] = {}
    for dash_id, d in dashboards.items():
        analysis = d.tree_app.analysis
        if analysis is None:
            all_sheets[dash_id] = {str(d.sheet.sheet_id): d.sheet}
        else:
            all_sheets[dash_id] = {
                str(s.sheet_id): s for s in analysis.sheets
            }
    listing: list[tuple[str, str]] = [
        (dash_id, d.title) for dash_id, d in dashboards.items()
    ]
    # Use the first dashboard's theme for the listing page — if any
    # server hosts dashboards with different themes the listing
    # picks the alphabetically-first one's. That edge case isn't
    # the design target (one L2 instance → one theme); flagging it
    # as a comment so a future multi-tenant story sees the seam.
    listing_theme = next(iter(dashboards.values())).theme

    async def index(_request: Request) -> RedirectResponse:
        # ``/`` is a convenience redirect; ``/dashboards`` is the
        # canonical list page. Status 302 (temporary) since which
        # dashboard a future multi-tenant home would land on
        # could shift per-user.
        return RedirectResponse("/dashboards", status_code=302)

    docs_url = "/docs/" if docs_dir is not None else None
    # Phase BS.3 — studio_enabled derived from "studio_routes spliced
    # in" since that's the upstream signal. cli/studio.py passes
    # routes (studio surface mounted); cli/dashboards.py doesn't.
    # Cfg.studio_enabled gates earlier in the CLI per BS.0 Lock 1.
    studio_enabled = studio_routes is not None

    def _render_top_nav(active_href: str | None) -> str:
        """Phase BS.3 — render the shared flat top-nav for any page.
        Caller passes the current URL path so the matching entry
        renders as active."""
        from recon_gen.common.html.render import (  # noqa: PLC0415
            build_top_nav_entries, emit_top_nav,
        )
        return emit_top_nav(
            entries=build_top_nav_entries(
                listing,
                studio_enabled=studio_enabled,
                docs_url=docs_url,
            ),
            active_href=active_href,
        )

    async def dashboards_list(_request: Request) -> HTMLResponse:
        return HTMLResponse(
            emit_dashboards_list(
                listing, theme=listing_theme, docs_url=docs_url,
                studio_enabled=studio_enabled,
                banner_text=banner_text,
            ),
        )

    async def dashboard_view(request: Request) -> Response:
        dash_id = request.path_params["dashboard_id"]
        served = dashboards.get(dash_id)
        if served is None:
            # Raise so the themed 404 handler renders the page,
            # not Starlette's default plain-text "Not Found" body.
            raise HTTPException(status_code=404)
        # Tab strip across the top — every analysis sheet becomes a tab.
        # Single-sheet dashboards get an empty tab strip (suppressed
        # by ``_render_sheet_tabs``).
        sheets = tuple(all_sheets[dash_id].values())
        # Y.2.app2.cde.l2ft-wiring.b — auto-derive per-sheet filter specs
        # from the tree when the dashboard didn't supply explicit ones.
        filter_specs = served.filter_specs or tuple(
            make_filter_specs_for_sheet(served.sheet),
        )
        # CQ.2.c — LinkedValues dropdowns carry empty <option> lists
        # by design; Tom Select's load callback fetches the seed page
        # on first focus. Identity passthrough kept until call sites
        # drop the call.
        filter_specs = _resolve_linked_options(filter_specs)
        # u.4.e.4 — ?param_<name>=<v> in the page URL (a drill that walked
        # an anchor, or a bookmark) pre-selects the matching widget so the
        # visuals' hx-include="#filter-form" load fetch is already narrowed.
        filter_specs = _apply_url_param_overrides(
            filter_specs, request.query_params,
        )
        # X.4.g.12.b — capture the current generation counter at render
        # time. The page's poller will compare against this baseline.
        from recon_gen.common.l2.deploy_pipeline import (  # noqa: PLC0415
            get_data_generation_id,
        )
        # BV.4.8.P1.1 — Dual-prefix Trainer Tour link carries
        # ``?prefix=<base>_v`` to point this page at the v overlay.
        # Threaded into the filter form as a hidden input so every
        # visual's hx-include="#filter-form" fetch carries it.
        url_prefix = request.query_params.get("prefix") or None
        return HTMLResponse(emit_html(
            served.tree_app, served.sheet,
            dashboard_id=dash_id, dev_log=dev_log,
            theme=served.theme,
            all_sheets=sheets,
            filter_specs=filter_specs,
            data_generation_id=get_data_generation_id(),
            top_nav=_render_top_nav(active_href=f"/dashboards/{dash_id}"),
            prefix_override=url_prefix,
            banner_text=banner_text,
        ))

    async def sheet_view(request: Request) -> Response:
        """X.2.e — render a specific sheet by id.

        Plain-anchor sheet tabs target this route. The dashboard's
        analysis must contain a sheet with the matching id; unknown
        ids 404 (themed via the same handler the dashboard route
        uses).
        """
        dash_id = request.path_params["dashboard_id"]
        served = dashboards.get(dash_id)
        if served is None:
            raise HTTPException(status_code=404)
        sheet_id = request.path_params["sheet_id"]
        sheet_for_dash = all_sheets[dash_id].get(sheet_id)
        if sheet_for_dash is None:
            raise HTTPException(status_code=404)
        sheets = tuple(all_sheets[dash_id].values())
        # Y.2.app2.cde.l2ft-wiring.b — per-sheet auto-derive (see dashboard_view).
        filter_specs = served.filter_specs or tuple(
            make_filter_specs_for_sheet(sheet_for_dash),
        )
        # CQ.2.c — see dashboard_view; identity passthrough.
        filter_specs = _resolve_linked_options(filter_specs)
        # u.4.e.4 — see dashboard_view; a ?param_<name>=<v> in the URL
        # pre-selects the matching widget so the load fetch is narrowed.
        filter_specs = _apply_url_param_overrides(
            filter_specs, request.query_params,
        )
        # X.4.g.12.b — same poller baseline as dashboard_view.
        from recon_gen.common.l2.deploy_pipeline import (  # noqa: PLC0415
            get_data_generation_id,
        )
        # BV.4.8.P1.1 — see dashboard_view.
        url_prefix = request.query_params.get("prefix") or None
        return HTMLResponse(emit_html(
            served.tree_app, sheet_for_dash,
            dashboard_id=dash_id, dev_log=dev_log,
            theme=served.theme,
            all_sheets=sheets,
            filter_specs=filter_specs,
            data_generation_id=get_data_generation_id(),
            top_nav=_render_top_nav(active_href=f"/dashboards/{dash_id}"),
            prefix_override=url_prefix,
            banner_text=banner_text,
        ))

    async def visual_data(request: Request) -> Response:
        # 404 on stale URLs — both ids must resolve. The visual_id
        # gets validated implicitly (the fetcher raises for
        # unknown ids; that's the per-fetcher contract).
        dash_id = str(request.path_params["dashboard_id"])
        served = dashboards.get(dash_id)
        if served is None:
            raise HTTPException(status_code=404)
        # X.2.e — any analysis sheet's visual is fetchable, not just
        # the served (default landing) sheet. The fetcher resolves
        # the visual_id; the sheet_id check protects against typos
        # in the URL pattern.
        sheet_id = str(request.path_params["sheet_id"])
        if sheet_id not in all_sheets[dash_id]:
            raise HTTPException(status_code=404)
        # X.2.o.3 — wrap path-extracted str into VisualId at the
        # route boundary so the fetcher sees the typed identifier
        # the DataFetcher contract requires. Path params come back
        # as ``Any`` from Starlette; ``str()`` narrows then
        # ``VisualId(...)`` brands.
        visual_id = VisualId(str(request.path_params["visual_id"]))
        # ``Mapping[str, list[str]]`` interface — built mutable for
        # the short construction window, then handed to the fetcher
        # contract that promises read-only access. ``multi_items()``
        # preserves repeated keys (``?param_pRail=A&param_pRail=B`` →
        # ``{"param_pRail": ["A", "B"]}``) so the SQL executor can
        # expand a multi-valued ``IN``-list (Y.2.app2.cde.multivalued);
        # single-valued params land as one-element lists.
        params: dict[str, list[str]] = {}
        for key, value in request.query_params.multi_items():
            params.setdefault(str(key), []).append(str(value))
        # X.2.n.5 — dispatch async fetchers directly so the asyncio
        # loop stays free across the SQL roundtrip; only sync stub
        # fetchers (tests + legacy _db_fetcher) get the threadpool
        # offload. The threadpool fallback keeps the contract
        # backward-compatible without forcing every test to become
        # async.
        if inspect.iscoroutinefunction(served.data_fetcher):
            data = await served.data_fetcher(visual_id, params)
        else:
            data = await run_in_threadpool(
                served.data_fetcher, visual_id, params,
            )
        # CH.5 (2026-06-08) — XLSX export branch. When `?format=xlsx`
        # the rendered table downloads as an Excel file with the
        # SAME columns + rows currently on screen (paginated/filtered
        # via the same SQL params). Currency columns format as Excel
        # currency right-aligned. Only meaningful for table-shaped
        # data (`shape_table` output); other shapes return 400.
        if str(request.query_params.get("format") or "") == "xlsx":
            xlsx_bytes = _emit_xlsx_workbook(data, visual_id)
            filename = f"{dash_id}-{sheet_id}-{visual_id}.xlsx"
            return Response(
                content=xlsx_bytes,
                media_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="{filename}"'
                    ),
                },
            )
        return HTMLResponse(
            # AA.B.5.followon.diag — pass URL params through so the
            # fragment stamps them as ``data-bound-params`` on the
            # script tag. Failure-capture ``dom.html`` then carries
            # per-visual evidence of what each fetch was queried with
            # (distinguishes "valid pick, no matching rows" from
            # "pick value never made it to the server").
            emit_visual_data_fragment(visual_id, data, url_params=params),
            headers={"Cache-Control": visual_data_cache_header},
        )

    async def handbook_page(request: Request) -> Response:
        """CN.5 — render a handbook page for the App2 ``?`` side panel.

        Route: ``GET /handbook/{handbook_path:path}``. The path comes
        from a Sheet's ``handbook_path: HandbookPath`` field (e.g.
        ``l1/drift``); resolved to ``docs/handbook/<path>.md`` on disk
        and rendered through the same markdown pipeline the rest of
        the App2 surface uses. Returns an HTML fragment ready for
        injection into the side panel container — no full-page chrome.

        404 if the file doesn't exist (e.g. a Sheet declared a path
        that wasn't authored yet) or if the path tries to escape the
        handbook directory.
        """
        from pathlib import Path as _Path  # noqa: PLC0415 — lazy
        import markdown as _md  # noqa: PLC0415 — lazy

        # server.py lives at src/recon_gen/common/html/server.py;
        # parents[0]=html, [1]=common, [2]=recon_gen, [3]=src, [4]=repo
        repo_root = _Path(__file__).resolve().parents[4]
        handbook_root = repo_root / "docs" / "handbook"
        raw_path = request.path_params["handbook_path"]
        # Resolve + ensure the final path is still inside handbook_root.
        # Defense against ``../../etc/passwd`` style traversal.
        candidate = (handbook_root / f"{raw_path}.md").resolve()
        try:
            candidate.relative_to(handbook_root.resolve())
        except ValueError:
            raise HTTPException(status_code=404)  # noqa: B904
        if not candidate.is_file():
            raise HTTPException(status_code=404)
        text = candidate.read_text(encoding="utf-8")
        html_body: str = _md.markdown(
            text,
            extensions=["fenced_code", "tables", "toc"],
        )
        # The side panel injects this as innerHTML; wrap in a
        # ``<article>`` so its prose-class styles can target the
        # handbook content specifically.
        wrapped = (
            "<article class=\"handbook-page prose prose-sm "
            "max-w-none\">"
            f"{html_body}"
            "</article>"
        )
        return HTMLResponse(wrapped)

    async def dropdown_options(request: Request) -> Response:
        """BR.1 — App2 cascade refresh endpoint.

        Returns the ``<option>`` HTML fragments for a cascading
        dropdown's source dataset, narrowed by the current form state.
        The rendered ``<select>`` carries ``hx-get`` against this URL +
        ``hx-trigger="change from:[name='param_<source>']"`` so HTMX
        swaps the option list whenever the source picker changes.

        Without this route, ``_resolve_linked_options`` only fires at
        page-load — picking the Role dropdown refreshes only visuals,
        the Account dropdown's options stay stale (the QS-side BO.1
        bug's App2 twin; see ``apps/l1_dashboard/app.py`` cascade fix).

        Preserves the user's current selection when it remains in the
        new option universe (re-emits a ``selected`` attribute);
        otherwise the swap clears it and the form's next fire narrows
        to "no selection" until the user picks again — exactly the
        QS cascade behavior.
        """
        dash_id = str(request.path_params["dashboard_id"])
        served = dashboards.get(dash_id)
        if served is None:
            raise HTTPException(status_code=404)
        sheet_id = str(request.path_params["sheet_id"])
        if sheet_id not in all_sheets[dash_id]:
            raise HTTPException(status_code=404)
        if served.options_search_fetcher is None:
            return HTMLResponse('<option value=""></option>')
        dataset_id = str(request.path_params["dataset"])
        column = str(request.path_params["column"])
        url_params = _query_params_as_multidict(request.query_params)
        # CQ.2.e — cascade route now calls the search fetcher with
        # query='' (seed page of the narrowed universe). Pre-CQ.2 it
        # called the silent LIMIT 2000 options_fetcher; same option
        # universe, narrower page, no truncation surprise.
        # CR.2 — cascade always passes ``query=""`` so the result's
        # ``truncated`` flag is structurally False; we ignore it
        # here (the HTML cascade has no banner surface — the JSON
        # typeahead does and surfaces it in the response payload).
        result = await served.options_search_fetcher(
            dataset_id, column, "", url_params,
        )
        opts = result.options
        # Preserve user's current pick if it survives the narrow. The
        # form's `param_<self>` key lives in the same query_params we
        # already parsed; look it up by walking the spec's URL key.
        # Self-name comes from the form params via the same `param_X`
        # convention every dropdown uses (the renderer's <select
        # name="param_X">), so we just check if any URL key matches
        # one of the new options.
        selected = ""
        for key, vals in url_params.items():
            if not key.startswith("param_"):
                continue
            for v in vals:
                if v and v in opts:
                    selected = v
                    break
            if selected:
                break
        parts = ['<option value=""></option>']
        for opt in opts:
            esc = html_escape(opt)
            if opt == selected:
                parts.append(f'<option value="{esc}" selected>{esc}</option>')
            else:
                parts.append(f'<option value="{esc}">{esc}</option>')
        return HTMLResponse("".join(parts))

    async def dropdown_search(request: Request) -> Response:
        """CQ.2.b — server-side typeahead JSON endpoint.

        Tom Select's ``load`` callback fires this on each typed
        keystroke (debounced via the built-in 300ms ``loadThrottle``).
        Empty ``q`` → seed page (top-N alphabetical, used by
        ``preload: 'focus'``); typed ``q`` → case-insensitive
        substring match.

        Response shape: ``{"options": [{"value": "...", "label": "..."}]}``
        — Tom Select's ``valueField: 'value'`` + ``labelField: 'label'``
        consume it. Capped at ``PICKER_PAGE_SIZE`` (100) server-side;
        the operator narrows further by typing more chars.

        Form-state passthrough: ``param_<name>`` query params thread
        through to ``execute_visual_sql_async`` so cascading source
        narrowing (e.g. Role narrows the candidate Account universe)
        still applies — same behavior as the HTML cascade endpoint,
        different transport.
        """
        dash_id = str(request.path_params["dashboard_id"])
        served = dashboards.get(dash_id)
        if served is None:
            raise HTTPException(status_code=404)
        sheet_id = str(request.path_params["sheet_id"])
        if sheet_id not in all_sheets[dash_id]:
            raise HTTPException(status_code=404)
        if served.options_search_fetcher is None:
            return JSONResponse({"options": []})
        dataset_id = str(request.path_params["dataset"])
        column = str(request.path_params["column"])
        query = str(request.query_params.get("q", ""))
        url_params = _query_params_as_multidict(request.query_params)
        # Drop the search ``q`` from the form-state multi-dict — it
        # rides the URL but isn't a ``param_<name>`` form field, and
        # ``collect_bind_params`` walks the SQL for ``:q`` separately
        # (the fetcher binds it via ``extra_binds``).
        url_params = {
            k: v for k, v in url_params.items() if k != "q"
        }
        # CR.2 — typeahead response now surfaces ``truncated`` so the
        # client UI can banner the silent-match-failure case (customer
        # typed > cap chars, fetcher trimmed). Pre-CR.2 the cap was a
        # hardcoded 100; CR.2 raises the default to 500 (operator-
        # tunable via ``RECON_GEN_PICKER_MAX_QUERY_LEN``) AND surfaces
        # the truncation event so a customer with realistic identifiers
        # gets a visible signal instead of an empty result with no clue.
        result = await served.options_search_fetcher(
            dataset_id, column, query, url_params,
        )
        return JSONResponse({
            "options": [{"value": v, "label": v} for v in result.options],
            "truncated": result.truncated,
        })

    async def log_event(request: Request) -> Response:
        try:
            payload = await request.json()
        except (json.JSONDecodeError, ValueError):
            payload = {"event": "dev-log:bad-json"}
        # Y.2.gate.c.11.app2-server-logs — route through Python's
        # logging module so the per-run log capture in the test
        # harness (`tests/e2e/_harness_html2.py`) lands these
        # browser-forwarded events alongside uvicorn's access log.
        # The ``DEV-LOG`` prefix preserves the grep-friendly shape
        # the previous stderr `print` had.
        _DEVLOG.info("DEV-LOG %s", json.dumps(payload))  # typing-smell: ignore[json-indent]: log line — compact one-line JSON is grep/jq-friendly
        return Response(status_code=204)

    # X.2.m — themed error pages for 4xx / 5xx. The handlers reuse
    # ``listing_theme`` so the error page inherits the per-dashboard
    # theme (the same picking convention as the ``/dashboards``
    # listing — the first dashboard's theme wins). A future
    # multi-tenant story that mixes themes per request would route
    # the per-request theme through here; flagged as a comment.
    async def not_found_handler(
        _request: Request, exc: Exception,
    ) -> Response:
        # Subtitle differs slightly when the URL pattern matched a
        # dashboard route vs. a generic path — but we can't always
        # tell from the exception alone (Starlette routes that
        # don't match at all also raise 404). Single message keeps
        # the contract simple: "the URL didn't resolve, here's
        # the way back."
        del exc
        return HTMLResponse(
            emit_error_page(
                status_code=404,
                headline="Not found",
                subtitle=(
                    "We couldn't find that dashboard or sheet. "
                    "Bookmarks may be stale; the link below goes "
                    "back to the dashboards list."
                ),
                theme=listing_theme,
                banner_text=banner_text,
            ),
            status_code=404,
        )

    # CS.8 — themed 503 for the CO.x PoolReleasedDuringRefresh window.
    # The studio's deploy pipeline brackets step_1_etl_hook by
    # releasing the dashboards' DuckDB pool root (so the subprocess
    # can hold the writer lock); any dashboard request that lands
    # during that bracket raises PoolReleasedDuringRefresh. Operators
    # used to see the framework default 500 + a stack trace; now they
    # see a calm "data refresh in progress, page will reload" with an
    # auto-refresh meta tag.
    async def pool_released_handler(
        _request: Request, exc: Exception,
    ) -> Response:
        del exc
        return HTMLResponse(
            emit_error_page(
                status_code=503,
                headline="Data refresh in progress",
                subtitle=(
                    "A Studio data refresh is in flight; the demo "
                    "DB is briefly unavailable. The page reloads "
                    "automatically in 5 seconds, or click below to "
                    "go to the dashboards listing now."
                ),
                theme=listing_theme,
                auto_reload_secs=5,
                banner_text=banner_text,
            ),
            status_code=503,
        )

    async def server_error_handler(
        _request: Request, exc: Exception,
    ) -> Response:
        # Dev-mode carries the traceback inside <details>; production
        # hides it. ``traceback.format_exception`` gives the same
        # shape Python prints to stderr when an exception goes
        # uncaught — easiest for an operator to recognize when the
        # page lands in a screenshot.
        if dev_log:
            tb_text = "".join(
                traceback.format_exception(type(exc), exc, exc.__traceback__)
            )
        else:
            tb_text = None
        return HTMLResponse(
            emit_error_page(
                status_code=500,
                headline="Something went wrong",
                subtitle=(
                    "We hit an error rendering this dashboard. Try "
                    "again, or contact your admin if it persists."
                ),
                traceback_text=tb_text,
                theme=listing_theme,
                banner_text=banner_text,
            ),
            status_code=500,
        )

    # X.4.g.12 — Step 5 reload counter endpoint. Open Dashboards pages
    # (well, future ones — JS poller lands with X.4.g.14) read this to
    # decide whether to reload after a Studio deploy bumped the counter.
    # Always mounted; safe under both Studio + Dashboards-only modes.
    async def data_generation_id_route(_request: Request) -> JSONResponse:
        from recon_gen.common.l2.deploy_pipeline import (  # noqa: PLC0415
            get_data_generation_id,
        )
        return JSONResponse({"data_generation_id": get_data_generation_id()})

    # Tailwind CSS lives next to this module in assets/; built by
    # ``.venv/bin/tailwindcss -i .../assets/input.css -o
    # .../assets/output.css``. Page shell links it as
    # ``/static/output.css``. Tracked in git so the spike runs
    # without forcing the user to build CSS first.
    assets_dir = Path(__file__).parent / "assets"

    routes: list[Route | Mount] = []
    if studio_routes is not None:
        # X.4.a.4 — Studio mount owns ``GET /`` (its landing page);
        # skip the Dashboards-only redirect. Splice Studio routes
        # first so a Studio-defined ``GET /`` wins on the route table.
        routes.extend(studio_routes)
    else:
        routes.append(Route("/", index, methods=["GET"]))
    routes += [
        Route("/dashboards", dashboards_list, methods=["GET"]),
        Route(
            "/dashboards/{dashboard_id}",
            dashboard_view, methods=["GET"],
        ),
        Route(
            "/dashboards/{dashboard_id}/sheets/{sheet_id}",
            sheet_view, methods=["GET"],
        ),
        Route(
            "/dashboards/{dashboard_id}/sheets/{sheet_id}"
            "/visuals/{visual_id}/data",
            visual_data,
            methods=["GET"],
        ),
        Route(
            "/dashboards/{dashboard_id}/sheets/{sheet_id}"
            "/dropdown-options/{dataset}/{column}",
            dropdown_options, methods=["GET"],
        ),
        # CQ.2.b — server-side typeahead JSON endpoint. Parallel to
        # the HTML cascade route above; different consumer (Tom
        # Select's ``load`` callback fetches JSON, vs. the cascade
        # route's HTMX HTML swap on sibling-control change).
        Route(
            "/dashboards/{dashboard_id}/sheets/{sheet_id}"
            "/dropdown-search/{dataset}/{column}",
            dropdown_search, methods=["GET"],
        ),
        # CY.5 — row-metadata side-panel fragment for tables wired
        # with ``Table.metadata_popup=True``. Stateless: the metadata
        # JSON travels as a query param sourced from the row payload
        # (CY.4 projects ``metadata`` on every row). Validates
        # dashboard / sheet / table-has-metadata-popup, parses the
        # JSON, renders the collapsible ``<details>`` tree.
        Route(
            "/dashboards/{dashboard_id}/sheets/{sheet_id}"
            "/rows/metadata",
            metadata_panel_route_factory(dashboards, all_sheets),
            methods=["GET"],
        ),
        # CN.5 — handbook page fetch for the App2 ``?`` side panel.
        # Path converter ``:path`` lets it match nested slugs like
        # ``l1/drift`` or ``_shared/app-info``.
        Route(
            "/handbook/{handbook_path:path}",
            handbook_page, methods=["GET"],
        ),
        Mount(
            "/static",
            app=StaticFiles(directory=str(assets_dir)),
            name="static",
        ),
        Route(
            "/data_generation_id",
            data_generation_id_route, methods=["GET"],
        ),
    ]
    if docs_dir is not None:
        # X.2.i — embed the *built* mkdocs handbook. ``html=True`` makes
        # ``/docs/`` → ``…/index.html`` and ``/docs/handbook/l1/`` →
        # ``…/handbook/l1/index.html`` (mkdocs's pretty-URL layout).
        # Mounting at ``/docs`` (not the site root) relies on the site
        # using relative internal links — mkdocs-material's default.
        routes.append(Mount(
            "/docs",
            app=StaticFiles(directory=str(docs_dir), html=True),
            name="docs",
        ))
    if dev_log:
        routes.append(Route("/log", log_event, methods=["POST"]))
    # exception_handlers maps status code (HTTPException) OR exception
    # class (everything else) → handler. 404 goes via status code so
    # it catches both raises from our route handlers AND Starlette's
    # own "no route matched" 404. Generic ``Exception`` catches any
    # uncaught throw from a fetcher / render path so production never
    # returns the framework default page.
    return Starlette(
        debug=False,
        routes=routes,
        exception_handlers={
            404: not_found_handler,
            # CS.8 — explicit handler for the CO.x typed exception so
            # the page renders as 503 + auto-reload instead of the
            # generic 500 + traceback served by server_error_handler.
            PoolReleasedDuringRefresh: pool_released_handler,
            Exception: server_error_handler,
        },
    )
