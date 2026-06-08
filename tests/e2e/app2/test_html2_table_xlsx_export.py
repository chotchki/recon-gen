# pyright: reportUnknownLambdaType=false, reportUnknownMemberType=false
"""CH.5 — App 2's "↓ XLSX" download matches what the HTML renders.

Browser round-trip: click the toolbar link above a Table visual, capture
the downloaded `.xlsx` bytes, parse with openpyxl, and assert columns +
row values match `App2Driver.table_rows()` (the same parsed page the
operator is looking at). Currency-typed cells also carry the right-
aligned currency number-format the unit test pins.

This is the X.4 offline-iteration parity gate for export: if a SQL
projection adds a column or changes a row, the XLSX changes too — no
silent divergence between the on-screen render and the download.

App2-only — QuickSight's own export-to-CSV/PDF stays under AWS's path.
"""

from __future__ import annotations

from io import BytesIO

import pytest


playwright_sync_api = pytest.importorskip("playwright.sync_api")  # noqa: F841
pytest.importorskip("openpyxl")

from openpyxl import load_workbook

from tests.e2e._drivers import App2Driver


_TABLE = "Account Balances"


def _open_showcase(d: App2Driver) -> None:
    d.open("smoke", sheet="Showcase")
    d.wait_loaded(_TABLE)


def test_xlsx_download_matches_rendered_table_rows() -> None:
    with App2Driver.smoke() as d:
        _open_showcase(d)

        rendered = d.table_rows(_TABLE)
        assert rendered, "smoke table should render at least one row"
        rendered_columns = list(rendered[0].keys())

        link = d.page.locator(".table-xlsx-download").first
        assert link.count() == 1, "↓ XLSX link should be present"

        with d.page.expect_download() as info:
            link.click()
        download = info.value
        path = download.path()
        assert path is not None, "download should resolve to a temp file"
        payload = path.read_bytes()
        assert payload[:2] == b"PK", "downloaded payload should be a zip (.xlsx)"

        workbook = load_workbook(BytesIO(payload))
        sheet = workbook.active
        assert sheet is not None

        header_row = [cell.value for cell in sheet[1]]
        assert header_row == rendered_columns, (
            f"XLSX header drift: {header_row} vs rendered {rendered_columns}"
        )

        # XLSX page is the same slice the HTML renders — same row count,
        # same first-column ordering. Cell values stringify because the
        # HTML parsing route strips currency formatting.
        xlsx_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        assert len(xlsx_rows) == len(rendered), (
            f"XLSX rendered {len(xlsx_rows)} rows vs HTML {len(rendered)}"
        )
        first_col = rendered_columns[0]
        for xrow, rrow in zip(xlsx_rows, rendered):
            assert str(xrow[0]) == rrow[first_col], (
                f"first-column drift: xlsx={xrow[0]!r} html={rrow[first_col]!r}"
            )


def test_xlsx_download_applies_currency_format_to_money_columns() -> None:
    with App2Driver.smoke() as d:
        _open_showcase(d)

        rendered = d.table_rows(_TABLE)
        assert rendered

        link = d.page.locator(".table-xlsx-download").first
        with d.page.expect_download() as info:
            link.click()
        path = info.value.path()
        assert path is not None
        payload = path.read_bytes()

        workbook = load_workbook(BytesIO(payload))
        sheet = workbook.active
        assert sheet is not None
        headers = [cell.value for cell in sheet[1]]

        money_columns = [
            i for i, h in enumerate(headers, start=1)
            if isinstance(h, str) and "balance" in h.lower()
        ]
        assert money_columns, (
            "smoke 'Account Balances' should carry at least one balance column"
        )

        for col_idx in money_columns:
            for row in range(2, min(sheet.max_row, 4) + 1):
                cell = sheet.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                # Currency format pinned in server._emit_xlsx_workbook;
                # right-aligned in the same place.
                assert "$" in (cell.number_format or ""), (
                    f"money cell at ({row},{col_idx}) lost currency fmt: "
                    f"{cell.number_format!r}"
                )
                assert cell.alignment.horizontal == "right", (
                    f"money cell at ({row},{col_idx}) should be right-aligned"
                )
